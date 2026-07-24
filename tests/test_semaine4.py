"""Tests des fonctionnalités avancées (semaine 4) : travaux, exports, duplication."""
import pytest
from openpyxl import load_workbook

from app.extensions import db
from app.models import LigneTravaux, Projet, Scenario

from helpers import creer_parcours_complet, inscrire_et_connecter


def test_travaux_detailles_synchronisent_budget(http, app):
    inscrire_et_connecter(http)
    _, projet_id, _ = creer_parcours_complet(http, app)

    http.post(f"/projets/{projet_id}/travaux",
              data={"libelle": "Cuisine équipée", "categorie": "Cuisine / Salle de bain",
                    "montant": "60000"})
    http.post(f"/projets/{projet_id}/travaux",
              data={"libelle": "Peinture", "categorie": "Décoration", "montant": "15000"})

    with app.app_context():
        projet = db.session.get(Projet, projet_id)
        # Le budget global (100 000 initialement) devient la somme des postes
        assert projet.budget_travaux == pytest.approx(75_000)
        ligne_id = LigneTravaux.query.filter_by(libelle="Peinture").one().id

    http.post(f"/projets/travaux/{ligne_id}/supprimer")
    with app.app_context():
        assert db.session.get(Projet, projet_id).budget_travaux == pytest.approx(60_000)


def test_export_pdf(http, app):
    inscrire_et_connecter(http)
    _, _, scenario_id = creer_parcours_complet(http, app)
    reponse = http.get(f"/exports/scenario/{scenario_id}/pdf")
    assert reponse.status_code == 200
    assert reponse.mimetype == "application/pdf"
    assert reponse.data[:5] == b"%PDF-"
    assert len(reponse.data) > 10_000  # document réel, pas une coquille vide


def test_export_excel(http, app):
    from io import BytesIO

    inscrire_et_connecter(http)
    _, _, scenario_id = creer_parcours_complet(http, app)
    reponse = http.get(f"/exports/scenario/{scenario_id}/excel")
    assert reponse.status_code == 200

    classeur = load_workbook(BytesIO(reponse.data))
    assert set(classeur.sheetnames) == {
        "Hypothèses", "Indicateurs", "Projection", "Amortissement"
    }
    projection = classeur["Projection"]
    # 20 années de projection + en-tête
    assert projection.max_row == 21
    amortissement = classeur["Amortissement"]
    assert amortissement.max_row == 241  # 240 mensualités + en-tête


def test_export_excel_cash_sans_amortissement(http, app):
    from io import BytesIO

    inscrire_et_connecter(http)
    _, projet_id, _ = creer_parcours_complet(http, app)
    http.post(
        f"/scenarios/nouveau/{projet_id}",
        data={"nom": "Cash", "mode": "cash", "apport": "0", "taux_interet": "0",
              "taux_assurance": "0", "duree_annees": "20", "horizon_annees": "15",
              "revalorisation_loyer_pct": "0", "revalorisation_bien_pct": "0",
              "frais_revente_pct": "0", "taux_actualisation": "3"},
    )
    with app.app_context():
        scenario_id = Scenario.query.filter_by(nom="Cash").one().id
    reponse = http.get(f"/exports/scenario/{scenario_id}/excel")
    classeur = load_workbook(BytesIO(reponse.data))
    assert "Amortissement" not in classeur.sheetnames


def test_duplication_scenario(http, app):
    inscrire_et_connecter(http)
    _, _, scenario_id = creer_parcours_complet(http, app)
    reponse = http.post(f"/scenarios/{scenario_id}/dupliquer")
    assert reponse.status_code == 302
    with app.app_context():
        copie = Scenario.query.filter(Scenario.nom.like("%variante%")).one()
        original = db.session.get(Scenario, scenario_id)
        assert copie.id != original.id
        assert copie.taux_interet == original.taux_interet
        assert copie.apport == original.apport


def test_exports_cloisonnes(http, app):
    inscrire_et_connecter(http)
    _, _, scenario_id = creer_parcours_complet(http, app)
    http.post("/auth/logout")
    inscrire_et_connecter(http, email="intrus@choubel.com", nom="Intrus")
    assert http.get(f"/exports/scenario/{scenario_id}/pdf").status_code == 404
    assert http.get(f"/exports/scenario/{scenario_id}/excel").status_code == 404
