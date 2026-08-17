"""Construction du classeur Excel d'un scénario (openpyxl)."""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ENTETE_FILL = PatternFill("solid", fgColor="005291")
ENTETE_FONT = Font(bold=True, color="FFFFFF")
FORMAT_MONTANT = "# ##0"
FORMAT_PCT = "0.00\\ \\%"


def _entetes(ws, titres: list[str]) -> None:
    ws.append(titres)
    for cellule in ws[1]:
        cellule.fill = ENTETE_FILL
        cellule.font = ENTETE_FONT
        cellule.alignment = Alignment(horizontal="center")
    for i, titre in enumerate(titres, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(len(titre) + 4, 16)


def _feuille_cles_valeurs(ws, lignes: list[tuple]) -> None:
    _entetes(ws, ["Paramètre", "Valeur"])
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22
    for cle, valeur, *format_ in lignes:
        ws.append([cle, valeur])
        if format_ and isinstance(valeur, (int, float)):
            ws.cell(row=ws.max_row, column=2).number_format = format_[0]


def construire_classeur(projet, scenario, r: dict, amortissement: list[dict]) -> BytesIO:
    devise = r["devise"]
    wb = Workbook()

    # ── Hypothèses ────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Hypothèses"
    en_credit = scenario.mode == "credit"
    _feuille_cles_valeurs(ws, [
        ("Client", projet.client.nom),
        ("Objectif du client", projet.client.brief.objectif_libelle if projet.client.brief else "—"),
        ("Horizon du client (années)", projet.client.brief.horizon_annees if projet.client.brief else "—"),
        ("Projet", projet.nom),
        ("Type d'opération", "Locatif" if projet.est_locatif else "Terrain / revente"),
        ("Étape du dossier", projet.statut_libelle),
        ("Délai de livraison (mois)", projet.delai_livraison_mois),
        ("Zone de marché", f"{projet.zone.nom} ({devise})"),
        (f"Prix du bien ({devise})", projet.prix_bien, FORMAT_MONTANT),
        ("Taux de frais d'acquisition (%)", projet.taux_frais_acquisition, FORMAT_PCT),
        (f"Budget travaux ({devise})", projet.budget_travaux, FORMAT_MONTANT),
        (f"Loyer mensuel ({devise})", projet.loyer_mensuel, FORMAT_MONTANT),
        (f"Charges copropriété annuelles ({devise})", projet.charges_copro_annuelles, FORMAT_MONTANT),
        (f"Assurance annuelle ({devise})", projet.assurance_annuelle, FORMAT_MONTANT),
        ("Frais de gestion (% du loyer)", projet.frais_gestion_pct, FORMAT_PCT),
        ("Vacance locative (% du loyer)", projet.vacance_pct, FORMAT_PCT),
        (f"Entretien annuel ({devise})", projet.entretien_annuel, FORMAT_MONTANT),
        (f"Taxe annuelle ({devise})", projet.taxe_annuelle, FORMAT_MONTANT),
        ("Taux d'imposition effectif (%)", projet.taux_imposition, FORMAT_PCT),
        ("Scénario", scenario.nom),
        ("Mode de financement", "Crédit" if en_credit else "Cash"),
        (f"Apport ({devise})", scenario.apport if en_credit else r["acquisition"]["cout_total"], FORMAT_MONTANT),
        ("Taux d'intérêt annuel (%)", scenario.taux_interet if en_credit else 0, FORMAT_PCT),
        ("Taux d'assurance annuel (%)", scenario.taux_assurance if en_credit else 0, FORMAT_PCT),
        ("Durée du prêt (années)", scenario.duree_annees if en_credit else "—"),
        ("Horizon de projection (années)", scenario.horizon_annees),
        ("Revalorisation du loyer (%/an)", scenario.revalorisation_loyer_pct, FORMAT_PCT),
        ("Revalorisation du bien (%/an)", scenario.revalorisation_bien_pct, FORMAT_PCT),
        ("Frais de revente (%)", scenario.frais_revente_pct, FORMAT_PCT),
        (f"Prix de revente saisi ({devise})", scenario.prix_revente or "—", FORMAT_MONTANT),
        ("Taux d'actualisation (%)", scenario.taux_actualisation, FORMAT_PCT),
    ])

    # ── Indicateurs ───────────────────────────────────────────────────────────
    # Ordre de lecture du cabinet : rendement net et cash-flow d'abord, valeur
    # créée ensuite, TRI et VAN en dernier.
    ws = wb.create_sheet("Indicateurs")
    tri = r["indicateurs"]["tri"]
    locatif = r.get("type_operation", "locatif") == "locatif"
    indicateurs: list[tuple] = []
    if locatif:
        indicateurs += [
            ("Rendement net (%)", r["rendements"]["net"], FORMAT_PCT),
            (f"Cash-flow mensuel ({devise})", r["indicateurs"]["cashflow_mensuel"], FORMAT_MONTANT),
            (f"Cash-flow annuel ({devise})", r["indicateurs"]["cashflow_annuel"], FORMAT_MONTANT),
            ("Rendement net-net (%)", r["rendements"]["net_net"], FORMAT_PCT),
            ("Rendement brut (%)", r["rendements"]["brut"], FORMAT_PCT),
        ]
    indicateurs += [
        (f"Valeur créée sur l'horizon ({devise})", r["indicateurs"]["valeur_creee"], FORMAT_MONTANT),
        (f"Valeur du bien à l'horizon ({devise})", r["revente"]["valeur_bien_horizon"], FORMAT_MONTANT),
        (f"Plus-value brute ({devise})", r["revente"]["plus_value_brute"], FORMAT_MONTANT),
        (f"Encaissé à la revente ({devise})", r["revente"]["revente_nette"], FORMAT_MONTANT),
        ("TRI (%)", tri if tri is not None else "Non calculable", FORMAT_PCT),
        (f"VAN ({devise})", r["indicateurs"]["van"], FORMAT_MONTANT),
        (f"Coût total d'acquisition ({devise})", r["acquisition"]["cout_total"], FORMAT_MONTANT),
        (f"dont frais d'acquisition ({devise})", r["acquisition"]["frais"], FORMAT_MONTANT),
        (f"Capital emprunté ({devise})", r["financement"]["capital_emprunte"], FORMAT_MONTANT),
        (f"Mensualité totale ({devise})", r["financement"]["mensualite_totale"], FORMAT_MONTANT),
        (f"Coût total du crédit ({devise})", r["financement"]["cout_total_credit"], FORMAT_MONTANT),
    ]
    _feuille_cles_valeurs(ws, indicateurs)

    # ── Projection annuelle ───────────────────────────────────────────────────
    ws = wb.create_sheet("Projection")
    _entetes(ws, ["Année", f"Loyer perçu ({devise})", f"Charges ({devise})",
                  f"Impôt ({devise})", f"Annuités ({devise})", f"Revente ({devise})",
                  f"Cash-flow ({devise})", f"Cumul ({devise})", f"CRD ({devise})"])
    for l in r["projection"]["lignes"]:
        ws.append([l["annee"], l["loyer"], l["charges"], l["impot"], l["annuite"],
                   l["revente"], l["cashflow"] + l["revente"], l["cumul"], l["crd"]])
    for ligne in ws.iter_rows(min_row=2, min_col=2):
        for cellule in ligne:
            cellule.number_format = FORMAT_MONTANT

    # ── Amortissement mensuel (mode crédit) ───────────────────────────────────
    if amortissement:
        ws = wb.create_sheet("Amortissement")
        _entetes(ws, ["Mois", f"Intérêts ({devise})",
                      f"Capital remboursé ({devise})", f"Capital restant dû ({devise})"])
        for l in amortissement:
            ws.append([l["mois"], l["interet"], l["capital_rembourse"], l["crd"]])
        for ligne in ws.iter_rows(min_row=2, min_col=2):
            for cellule in ligne:
                cellule.number_format = FORMAT_MONTANT

    flux = BytesIO()
    wb.save(flux)
    flux.seek(0)
    return flux
